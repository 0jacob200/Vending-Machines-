import constants as const
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Side
import subprocess
import sys
import os


def make_date_format(date):
    date = date.split("-")
    if len(date) == 3:
        day = int(date[2])
    elif len(date) == 2:
        day = 1
    else:
        return False
    try:
        date[0] = int(date[0])
        date[1] = int(date[1])
    except ValueError:
        return False
    if int(date[0]) < 1890 or int(date[0]) > 2190:
        return False
    try:
        date = datetime.date(date[0], date[1], day).strftime("%Y-%m")
    except ValueError:
        return False
    return date


def income_and_outcome_calculation(number_bank, card_num, customer_date):
    customer_date = make_date_format(customer_date)
    data_sms = []
    received_money = 0
    spent_money = 0
    messages = open("SMS_data.txt", "r")
    for sms in messages:
        sms = sms.split(" ")
        if sms[0] == const.BANKS_PHONENUMBER[number_bank]:
            date_sms = make_date_format(sms[1])
            card_to_check = sms[const.INDEX_OF_SMS["card"][number_bank]][:5]
            if customer_date == date_sms and card_to_check == card_num:
                sms = split_operation_in_sms(sms)
                data_sms.append(sms)
                if const.OPERATION_MEANS["received"][number_bank] in sms:
                    received_money += float(sms[const.INDEX_OF_SMS["sum"][number_bank]])
                elif const.OPERATION_MEANS["spent"][number_bank] in sms:
                    spent_money += float(sms[const.INDEX_OF_SMS["sum"][number_bank]])
    delta = received_money - spent_money
    return data_sms, received_money, spent_money, delta


def take_card_and_balance(number_bank):
    card_num_list = []
    balance_list = []
    messages = open("SMS_data.txt", "r")
    for sms in messages:
        sms = sms.split(" ")
        sms = split_operation_in_sms(sms)
        if sms[0] == const.BANKS_PHONENUMBER[number_bank]:
            card = sms[const.INDEX_OF_SMS["card"][number_bank]]
            card = card[:5]
            if card not in card_num_list:
                card_num_list.append(card)
                balance_list.append([])
            index_card = card_num_list.index(card)
            balance_list[index_card].append(sms[const.INDEX_OF_SMS["balance"][number_bank]])
    counter1 = 0
    new_balance_list = []
    while counter1 != len(card_num_list):
        new_balance_list.append(float(balance_list[counter1].pop()))
        counter1 += 1
    return card_num_list, new_balance_list


def total_per_month(date, card_list):
    counter3 = 0
    received_total = 0
    spent_total = 0
    delta_total = 0
    while counter3 != len(const.BANKS_NAME):
        for card in card_list[counter3]:
            data_sms, received, spent, delta = income_and_outcome_calculation(counter3, card, date)
            received_total += received
            spent_total += spent
            delta_total += delta
        counter3 += 1
    return received_total, spent_total, delta_total


def split_operation_in_sms(sms):
    for bank_phone_num in const.BANKS_PHONENUMBER:
        if sms[0] == bank_phone_num:
            index_bank = const.BANKS_PHONENUMBER.index(bank_phone_num)
            if const.SPLIT_OPERATION_NAME[index_bank] is True:
                operation_and_sum = sms[const.INDEX_OF_SMS["operation"][index_bank]]
                operation = operation_and_sum[0]
                sum_of_operation = operation_and_sum[1: len(operation_and_sum)]
                sms.insert(const.INDEX_OF_SMS["operation"][index_bank], operation)
                sms.insert(const.INDEX_OF_SMS["sum"][index_bank], sum_of_operation)
                sms.pop(const.INDEX_OF_SMS["sum"][index_bank] + 1)
    return sms


def creating_xls_file(data_set, date, bank, card):
    doc = Workbook()
    sheet = doc.active
    sheet.append(const.SHUPKA)
    total_received = 0
    total_spent = 0
    for sms in data_set:
        if const.BANKS_PHONENUMBER[bank] == sms[0]:
            operation = ''
            sum_of_operation = sms[const.INDEX_OF_SMS["sum"][bank]]
            if const.OPERATION_MEANS["received"][bank] in sms:
                operation = '+'
                total_received += float(sum_of_operation)
            elif const.OPERATION_MEANS["spent"][bank] in sms:
                operation = '-'
                total_spent += float(sum_of_operation)
            else:
                print('something wrong')
            operation_str = f'{operation}{float(sms[const.INDEX_OF_SMS["sum"][bank]])} {const.CURRENCY}'
            row = [sms[1], sms[2], const.BANKS_NAME[bank], sms[const.INDEX_OF_SMS["card"][bank]][:5], operation_str]
            sheet.append(row)
    total_sum = total_received - total_spent
    total_row_1 = [const.TOTAL, '', '', const.RECEIVED, f"{total_received} {const.CURRENCY}"]
    total_row_2 = ['', '', '', const.SPENT, f"{total_spent} {const.CURRENCY}"]
    total_row_3 = ['', '', '', const.DELTA, f"{total_sum} {const.CURRENCY}"]
    sheet.append(total_row_1)
    sheet.append(total_row_2)
    sheet.append(total_row_3)
    border = Border(left=Side(border_style='thin',
                              color='FF000000'),
                    right=Side(border_style='thin',
                               color='FF000000'),
                    top=Side(border_style='thin',
                             color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                 color='FF000000'),
                    vertical=Side(border_style='thin',
                                  color='FF000000'),
                    horizontal=Side(border_style='thin',
                                    color='FF000000')
                    )
    align_center = Alignment(horizontal='center',
                             vertical='bottom',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)
    for cell_1 in sheet[f'A1:E{len(data_set) + 4}']:
        for cell in cell_1:
            sheet[cell.coordinate].border = border
            sheet[cell.coordinate].alignment = align_center
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        # python 3.7.2 has some problems with coordinates of excel
        if sys.version[:5] == '3.7.2' or "3.7.4":
            if col == 1:
                col = 'A'
            elif col == 2:
                col = 'B'
            elif col == 3:
                col = 'C'
            elif col == 4:
                col = 'D'
            elif col == 5:
                col = 'E'
            else:
                print('something wrong')
        sheet.column_dimensions[col].width = float(value)
    fill = PatternFill(fill_type='solid',
                       start_color='c1c1c1',
                       end_color='c2c2c2')
    sheet['A1'].fill = fill
    sheet['B1'].fill = fill
    sheet['C1'].fill = fill
    sheet['D1'].fill = fill
    sheet['E1'].fill = fill
    counter = len(data_set) + 4
    while counter != len(data_set) + 1:
        sheet[f'A{counter}'].fill = fill
        counter -= 1
    counter = len(data_set) + 4
    while counter != len(data_set) + 1:
        sheet[f'B{counter}'].fill = fill
        counter -= 1
    counter = len(data_set) + 4
    while counter != len(data_set) + 1:
        sheet[f'C{counter}'].fill = fill
        counter -= 1
    counter = len(data_set) + 4
    while counter != len(data_set) + 1:
        sheet[f'D{counter}'].fill = fill
        counter -= 1
    counter = len(data_set) + 4
    while counter != len(data_set) + 1:
        sheet[f'E{counter}'].fill = fill
        counter -= 1
    path_doc = f"report_{date}_card{card}.xlsx"
    doc.save(path_doc)
    open_file(path_doc)


def open_file(filename):
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, filename])


def creating_xl_for_all(date_cust):
    doc = Workbook()
    sheet = doc.active
    sheet.append(const.SHUPKA)
    total_received = 0
    total_spent = 0
    num_sms = 0
    file = open('SMS_data.txt', "r")
    for sms in file:
        sms = sms.split(' ')
        date_sms = make_date_format(sms[1])
        if date_cust == date_sms:
            ph_num_sms = sms[0]
            if ph_num_sms in const.BANKS_PHONENUMBER:
                index_bank = const.BANKS_PHONENUMBER.index(ph_num_sms)
                operation = ''
                sms = split_operation_in_sms(sms)
                sum_of_operation = sms[const.INDEX_OF_SMS["sum"][index_bank]]
                if const.OPERATION_MEANS["received"][index_bank] in sms:
                    operation = '+'
                    total_received += float(sum_of_operation)
                elif const.OPERATION_MEANS["spent"][index_bank] in sms:
                    operation = '-'
                    total_spent += float(sum_of_operation)
                else:
                    print('something wrong')
                operation_str = f'{operation}{float(sms[const.INDEX_OF_SMS["sum"][index_bank]])} {const.CURRENCY}'
                row = [sms[1], sms[2], const.BANKS_NAME[index_bank],
                       sms[const.INDEX_OF_SMS["card"][index_bank]][:5], operation_str]
                sheet.append(row)
                num_sms += 1
    total_sum = total_received - total_spent
    total_row_1 = [const.TOTAL, '', '', const.RECEIVED, f"{total_received} {const.CURRENCY}"]
    total_row_2 = ['', '', '', const.SPENT, f"{total_spent} {const.CURRENCY}"]
    total_row_3 = ['', '', '', const.DELTA, f"{total_sum} {const.CURRENCY}"]
    sheet.append(total_row_1)
    sheet.append(total_row_2)
    sheet.append(total_row_3)
    border = Border(left=Side(border_style='thin',
                              color='FF000000'),
                    right=Side(border_style='thin',
                               color='FF000000'),
                    top=Side(border_style='thin',
                             color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                 color='FF000000'),
                    vertical=Side(border_style='thin',
                                  color='FF000000'),
                    horizontal=Side(border_style='thin',
                                    color='FF000000')
                    )
    align_center = Alignment(horizontal='center',
                             vertical='bottom',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)
    for cell_1 in sheet[f'A1:E{num_sms + 4}']:
        for cell in cell_1:
            sheet[cell.coordinate].border = border
            sheet[cell.coordinate].alignment = align_center
    dims = {}
    list_column = ['A', 'B', 'C', 'D', 'E']
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
    for col, value in dims.items():
        # python 3.7.2 has some problems with coordinates of excel
        if sys.version[:5] == '3.7.2':
            col = list_column[col - 1]
        sheet.column_dimensions[col].width = float(value)
    fill = PatternFill(fill_type='solid',
                       start_color='c1c1c1',
                       end_color='c2c2c2')
    for col in list_column:
        sheet[f'{col}1'].fill = fill
        counter = num_sms + 4
        while counter != num_sms + 1:
            sheet[f'{col}{counter}'].fill = fill
            counter -= 1
    path_doc = f"report_{date_cust}.xlsx"
    doc.save(path_doc)
    open_file(path_doc)


# date = make_date_format('1992-05')
# creating_xl_for_all(date)
# data_sms, received_money, spent_money, delta = income_and_outcome_calculation(0, '*6677', '2019-05')
# creating_xls_file(data_sms, '2019-05', 0, '*6677')
# income_and_outcome_calculation(0, "*6678", "2019-04")
# a1, a2 = take_card_and_balance(0)
# print(f"{a1} {a2}")
